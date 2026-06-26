#!/usr/bin/env python3
"""
Generate program/overview.html and program/full-schedule.html
from the Excel Program sheet.
"""
import openpyxl
from datetime import datetime, timedelta
import html as html_lib
import re
import unicodedata
from urllib.parse import quote

# ─────────────────────────────────────────────
# 1. Static content — Keynote / Workshop / Tutorial details
# ─────────────────────────────────────────────

KEYNOTE_DETAILS = {
    'katrina': {
        'display_name': 'Katrina Ligett',
        'affiliation': 'The Hebrew University of Jerusalem, Israel',
        'title': 'Data Degrades with Use',
        'abstract': (
            'We often treat data as an infinitely reusable resource: a single dataset can support many '
            'analyses, train multiple models, and be shared widely without apparent cost. This talk argues '
            'that in important ways, data is not endlessly reusable. Instead, in certain contexts, data '
            'behaves like a consumable resource that degrades with use. The clearest example arises in the '
            'presence of privacy concerns. Fundamental results show that any informative public analysis of '
            'personal data inevitably leaks some information about the underlying individuals, and that these '
            'privacy losses accumulate across repeated uses of the same or overlapping datasets. If some level '
            'of privacy is to be preserved, this imposes intrinsic limits on how many times data can be used. '
            'In joint work under submission, we connect this perspective to the mosaic effect from legal '
            'scholarship, arguing that privacy risks arise not only from combining data pieces, but also from '
            'combining seemingly innocuous data uses. Data can also degrade with use even when privacy is not '
            'at stake. A line of work on adaptive data analysis shows that repeatedly querying the same '
            'dataset can lead to overfitting: results that appear valid on the dataset but fail to generalize '
            'to the underlying distribution, even when the dataset is very large. Recognizing data degradation '
            'opens a range of research directions, including systems for tracking and budgeting data use, '
            'algorithmic techniques to mitigate degradation, the role of synthetic data and data curators, '
            'and new models of non-worst-case adaptive computation.'
        ),
        'bio': (
            'Katrina Ligett is a Professor in the School of Computer Science and Engineering at The Hebrew '
            'University of Jerusalem, where she is also the director of the interdisciplinary Federmann '
            'Center for the Study of Rationality. Before joining Hebrew University, she was faculty in '
            'computer science and economics at Caltech. Her primary research interests are in data privacy, '
            'algorithmic fairness, machine learning theory, and algorithmic game theory. She received her PhD '
            'from Carnegie Mellon University in 2009 and did her postdoc at Cornell University. She is a '
            'recipient of an ERC grant, an NSF CAREER award, and a Microsoft Faculty Fellowship.'
        ),
    },
    'mounia': {
        'display_name': 'Mounia Lalmas',
        'affiliation': 'Spotify, UK',
        'title': 'Building AI-Driven Web Experiences at Scale',
        'abstract': (
            'AI is no longer just a component of Web systems; it is increasingly shaping the experiences '
            'users have online. From search and recommendation to conversational and generative interfaces, '
            'AI is redefining how people interact with content at Web scale. In this keynote, I reflect on '
            'how recent advances in AI, including deep learning and generative models, are reshaping the '
            'design space of Web technologies. Drawing on insights from developing AI-driven systems at '
            'Spotify, I discuss how search and recommendation are evolving into interactive, intent-aware '
            'experiences that support exploration and discovery. The talk highlights emerging system paradigms '
            'and research questions around building such experiences at scale, and reflects on the implications '
            'for the design of future Web systems and interactions.'
        ),
        'bio': (
            'Mounia Lalmas is Senior Director of Research at Spotify, where she leads Tech Research in '
            'Personalisation, focusing on search, recommendation, and discovery systems at scale. Her work '
            'spans information retrieval, recommender systems, user engagement, and the application of modern '
            'AI techniques to large-scale online platforms. She previously held senior roles including Director '
            'of Research at Yahoo and Professor of Information Retrieval at Queen Mary University of London. '
            'Mounia is an Honorary Professor at University College London and a Distinguished Research Fellow '
            'at the University of Amsterdam. She has co-chaired major conferences including SIGIR, WWW, WSDM, '
            'and CIKM, and authored over 260 publications.'
        ),
    },
    'alistair': {
        'display_name': 'Alistair Moffat',
        'affiliation': 'University of Melbourne, Australia',
        'title': '(Everything You Never Knew You Needed To Know About) Rank-Biased Measurement For Web Search',
        'abstract': (
            'In information retrieval and web search we measure how good search engines are at ordering '
            'answers to user queries, how close two ranked lists are to each other, how good LLMs are at '
            're-ranking sets of candidate documents, and how close generated answer sentences are to the ideal '
            'output. This talk begins by motivating the top-weighted measurements that arise when ordered '
            'sequences are involved, including reviewing the rank-biased precision and rank-biased overlap '
            'measurements proposed in 2008 and 2010. The second part then presents recent work unifying the '
            'two previous rank-biased approaches as elements in a larger framework that exposes a third '
            'rank-biased measurement, rank-biased recall. Finally, in the third part, new ways in which the '
            'degree of top-weighting bias can be controlled are described, allowing practitioners and '
            'researchers to better define their measurement goals and more precisely target their experiments.'
        ),
        'bio': (
            'Professor Alistair Moffat is now in his 40th year as a faculty member at the University of '
            'Melbourne. Early work included coauthorship of the book "Managing Gigabytes: Compressing and '
            'Indexing Documents and Images" (1994, 1999), and development of innovative mechanisms for '
            'implementing ranked queries via compressed inverted indexes. Most recently Alistair has been '
            'focused on IR evaluation, including models for user query formulation and result perusal, and '
            'rank-biased measurement. Alistair was inducted into the SIGIR Academy in 2021, and recently '
            'became a Fellow of the ACM for his contributions to the implementation and evaluation of '
            'search systems.'
        ),
    },
    'pascale': {
        'display_name': 'Pascale Fung',
        'affiliation': 'AMI Labs & The Hong Kong University of Science & Technology',
        'title': 'Towards AI That Understands the Human World',
        'abstract': (
            'AI has reached a turning point. Systems can now perceive, generate, and act in language and '
            'image across digital platforms at unprecedented scale. Yet as AI moves from tools to '
            'collaborators—embedded in decision-making, institutions, and everyday life—a new requirement '
            'becomes unavoidable: AI must understand the world the way humans inhabit it. This talk '
            'introduces Cognitive World Modeling as the next phase of AI development. It unifies physical '
            'world modeling—time, space, causality, action—with mental world modeling—goals, beliefs, '
            'intentions, emotions, and social norms—into a single, persistent representation of reality as '
            'experienced by humans. Together, these models allow AI systems not only to predict outcomes, '
            'but to reason about meaning, context, and consequence. Alignment and trust emerge not as post '
            'hoc constraints, but as properties of systems that maintain accurate, evolving models of both '
            'the external world and the humans within it.'
        ),
        'bio': (
            'Pascale Fung\'s long-term research background is in multimodal interactive systems including '
            'audio, speech, text and video. She is the Co-founder and Chief Research & Innovation Officer at '
            'AMI Labs. She was previously the Senior Director of AI Research at Meta-FAIR, leading research '
            'on embodied AI agents. She is also a Chair Professor of ECE at The Hong Kong University of '
            'Science & Technology (HKUST). She is a Fellow of the ACL, AAAI, IEEE, and ISCA for her '
            'significant contribution to human-machine interactions.'
        ),
    },
}


def keynote_lookup(event_text):
    if event_text is None:
        return None
    et = event_text.lower()
    for key in KEYNOTE_DETAILS:
        if key in et:
            return KEYNOTE_DETAILS[key]
    return None


WORKSHOP_DETAILS = {
    'AAIMVT': {
        'full_name': '1st Workshop on Applied AI and Multimodal Visualization Technologies',
        'organizers': 'Cesar Sanin (Australian Institute of Higher Education / University of New England), Edward Szczerbicki (University of Newcastle / Gdansk University of Technology), Md Rafiqul Islam (Charles Darwin University)',
        'url': 'https://rafiqulislamcse24.wixsite.com/aaimvt-26',
        'abstract': 'A full-day interactive workshop exploring how applied AI and multimodal visualization technologies can enhance knowledge representation, decision-making, and human-machine collaboration. Topics include cutting-edge research at the intersection of AI and multimodal visualization, interdisciplinary dialogue between researchers and practitioners, and methodologies to improve human decision-making through multimodal data representation.',
    },
    'AiOfAi': {
        'full_name': 'The Workshop on Adverse Impacts and Collateral Effects of Artificial Intelligence Technologies',
        'organizers': 'Esma Aimeur (Universite de Montreal), Rim Ben Salem (Polytechnique Montreal), Dorsaf Sallami (Universite de Montreal), Julita Vassileva (University of Saskatchewan), Nora Boulahia-Cuppens (Polytechnique Montreal)',
        'url': 'https://sites.google.com/view/aiofai-2026/home',
        'abstract': 'AiOfAi highlights the double-edged nature of AI in the digital age, examining how it can be exploited to undermine trust, privacy, and integrity, while also serving as a foundation for more secure, ethical, and resilient digital ecosystems. The workshop discusses the societal impact of widespread AI adoption, ethical and legal frameworks for responsible AI deployment, and emerging approaches in cybersecurity and fairness.',
    },
    'ALTARS': {
        'full_name': '4th Workshop on Augmented Intelligence in Technology-Assisted Review Systems (ALTARS 2026)',
        'organizers': 'Giorgio Maria Di Nunzio (University of Padova, Italy), Evangelos Kanoulas (University of Amsterdam), Prasenjit Majumder (DAIICT, Gandhinagar, India)',
        'url': 'https://altars2026.dei.unipd.it/',
        'abstract': 'ALTARS 2026 explores recent advances and open challenges in Technology-Assisted Review (TAR) systems for large-scale, high-recall retrieval across the Web. Topics include intelligent retrieval, human-in-the-loop learning, explainable and responsible AI, and the integration of LLMs and knowledge graphs into review workflows across legal, scientific, and web-scale domains.',
    },
    "BeyondFacts'26": {
        'full_name': '6th International Workshop on Computational Methods for Online Discourse Analysis',
        'organizers': 'Stefan Dietze (Heinrich-Heine-University Dusseldorf & GESIS), Dimitar Dimitrov (GESIS), Pavlos Fafalios (Technical University of Crete & FORTH-ICS), Konstantin Todorov (University of Montpellier / LIRMM / CNRS)',
        'url': 'https://beyondfacts2026.wordpress.com/',
        'abstract': 'This workshop strengthens relations between knowledge representation and NLP communities, providing a forum for works on modeling, extraction and analysis of online discourse. It addresses the need for shared understanding of discourse data—claims, arguments, stances, and veracity—using methods from machine learning, NLP, large language models and Web data mining.',
    },
    'DHOW-MiLLA': {
        'full_name': 'Joint Workshop on Diffusion of Harmful Content on Online Web and Countering Misinformation in the Age of LLMs and Agents',
        'organizers': 'Thomas Mandl (University of Hildesheim), Haiming Liu (University of Southampton), Gautam Kishore Shahi (University of Duisburg-Essen), Amit Kumar Jaiswal (IIT BHU Varanasi), and others',
        'url': 'https://dhow-workshop.github.io/2026/',
        'abstract': 'DHOW-MiLLA consolidates research on harmful content diffusion and misinformation under one umbrella. With LLMs and agent-based AI systems creating a dual-use paradigm, this workshop focuses on cross-platform, multilingual solutions that mitigate modern misinformation while harnessing AI capabilities for verification, fact-checking, and detection of deepfakes, propaganda, and multimodal disinformation.',
    },
    'FAAW': {
        'full_name': 'International Workshop on Foundations and Architectures for the Agentic Web',
        'organizers': 'Abderrahmane Maaradji (University of Doha for Science and Technology), Boualem Benatallah (Dublin City University), Fatma Outay (Zayed University, UAE), Ramesh Raskar (MIT Media Lab), Pradyumna Chari (MIT Media Lab), and others',
        'url': 'https://faaw.univ-tours.fr/',
        'abstract': 'The Agentic Web is emerging as billions of AI agents discover, communicate, and coordinate across the open Web. This workshop covers web-native building blocks: agent registries, identity and credentials (DIDs/VCs), authorization (OAuth 2.0), discovery (DNS-SD), and federation patterns, as well as economic mechanisms (reputation, knowledge pricing) and societal coordination (governance, accountability).',
    },
    'FL@FM': {
        'full_name': 'International Workshop on Federated Foundation Models for the Web 2026',
        'organizers': 'Irwin King (The Chinese University of Hong Kong), Guodong Long (University of Technology Sydney), Zenglin Xu (Fudan University), Han Yu (Nanyang Technological University), Xiaoli Tang (Nanyang Technological University)',
        'url': 'https://federated-learning.org/fl@fm-www-2026/',
        'abstract': 'With foundation models becoming the norm in ML development, federated learning (FL) becomes crucial for privacy-preserving and distributed learning at scale. This workshop provides a platform for researchers and industry professionals to discuss latest advancements in FL methods for foundation models, enabling efficient training while safeguarding sensitive data.',
    },
    'GLOW': {
        'full_name': 'Graph-enhanced LLMs for Trustworthy Web Data Management',
        'organizers': 'Gianluca Bonifazi (Marche Polytechnic University), Stefano Cirillo (University of Salerno), Eliana Pastor (Polytechnic University of Turin), Luca Virgili (Marche Polytechnic University)',
        'url': 'https://glow-workshop.github.io/www2026/',
        'abstract': 'This workshop explores synergies between LLMs and graph-based knowledge representations (knowledge graphs, property graphs) to build trustworthy data-driven Web applications. LLMs generate fluent responses but often struggle with factuality, bias, and hallucinations. Graphs provide structured, interconnected representations that can serve as grounding and validation layers for LLM-based systems.',
    },
    'HCRS': {
        'full_name': 'The 2nd Workshop on Human-Centered Recommender Systems',
        'organizers': 'Kaike Zhang (University of Chinese Academy of Sciences), Jiakai Tang (Renmin University of China), Julian McAuley (University of California, San Diego), Lina Yao (CSIRO Data61, Australia), and others',
        'url': 'https://hcrec.github.io/',
        'abstract': 'HCRS calls for a paradigm shift from optimizing engagement toward designing recommender systems that truly understand, involve, and benefit people. Centered around Human Understanding, Human Involvement, and Human Impact, the workshop covers topics from LLM-based interactive recommenders to societal welfare optimization and responsible recommendation research.',
    },
    'LARS': {
        'full_name': 'LLM & Agents for Recommendation Systems',
        'organizers': 'Keerthi Gopalakrishnan (Walmart Global Tech), Qi Xu (Meta AI), Aysenur Inan (Walmart Global Tech), Zhigang Hua (Meta AI), Shuang Yang (Meta AI), Luyi Ma (Walmart Global Tech)',
        'url': 'https://llmandagents4recsys.github.io/',
        'abstract': 'Recommendation systems are undergoing a major shift from traditional centralized pipelines to agentic ecosystems that can plan, reason, negotiate, and interact across the entire journey of discovery, personalization, and fulfillment. This workshop explores architectures, evaluation, trust, fairness, and real-world deployments to shape the next generation of adaptive, explainable recommendation ecosystems.',
    },
    'MM4SG': {
        'full_name': 'Fourth International Workshop on Multimodal Content Analysis for Social Good',
        'organizers': 'Usman Naseem (Macquarie University), Surendrabikram Thapa (Virginia Tech), Roy Ka-Wei Lee (Singapore University of Technology and Design), Mehwish Nasim (University of Western Australia)',
        'url': 'https://sites.google.com/view/mm4sg-webconf26',
        'abstract': 'MM4SG addresses the challenge of moderating multimodal content (memes, text-embedded images) on social platforms. The workshop brings together researchers from NLP, machine learning, computational social science, and ethics to explore innovative solutions for content moderation, sharing cutting-edge research on multimodal content analysis techniques.',
    },
    'PromptEng': {
        'full_name': 'Third International Workshop on Prompt Engineering for Pre-Trained Language Models',
        'organizers': 'Damien Graux (EcoVadis), Sebastien Montella (Huawei Technologies Ltd.), Hajira Jabeen (UniKlinik Cologne)',
        'url': 'https://prompteng-ws.github.io/2026/',
        'abstract': 'This workshop gathers practitioners to exchange about good practices, optimizations, results and novel paradigms for designing efficient prompts and context-building to make use of LLMs. Since LLM performances are highly dependent on the exact phrasing used in prompts, the workshop focuses on fail-retry strategies, prompt optimization, and novel prompting paradigms.',
    },
    'R2CASS': {
        'full_name': 'The 2nd International Workshop on Social Science Meets Web Data: Reproducible and Reusable Computational Approaches',
        'organizers': 'Fakhri Momeni (GESIS), Arnim Bleier (GESIS), Danilo Dessi (University of Sharjah, UAE), Muhammad Taimoor Khan (GESIS)',
        'url': 'https://sites.google.com/view/r2cass/home',
        'abstract': 'R2CASS advances computational reproducibility in social science, which relies on digital behavioral data from social media platforms. It brings together computer scientists, social scientists, and policy makers to improve the credibility and reproducibility of computational social science research. Features a hands-on session on the Methods Hub platform for computational reproducibility.',
    },
    'SemTech': {
        'full_name': '4th International Workshop on AI and Semantic Technologies for the Scientific, Technical, and Legal Web',
        'organizers': 'Rima Dessi (Higher College of Technologies, UAE), Jeenu Joy (FIZ-Karlsruhe), Danilo Dessi (University of Sharjah, UAE), Francesco Osborne (The Open University, UK), Hidir Aras (FIZ-Karlsruhe)',
        'url': 'https://semtech4stld.github.io/',
        'abstract': 'SemTech 2026 focuses on methods combining Semantic Web technologies, NLP, LLMs, and other AI to model knowledge across scientific, technical, and legal domains. The workshop invites research on knowledge graph creation, semantic annotation, LLM-KG hybrid reasoning, and trustworthy AI pipelines for scientific, patent, and legal Web content.',
    },
    'TempWeb': {
        'full_name': 'The 16th Temporal Web Analytics Workshop',
        'organizers': 'Marc Spaniol (University of Caen Normandy), Omar Alonso (Amazon), Ricardo Baeza-Yates (KTH, Royal Institute of Technology Stockholm)',
        'url': 'https://temporalweb.net/',
        'abstract': 'TempWeb provides a venue for researchers across all domains where the temporal dimension opens new challenges and possibilities. The workshop focuses on investigating infrastructures, scalable methods, and innovative software for aggregating, querying, and analyzing heterogeneous temporal data at Internet scale.',
    },
    'TrustFM': {
        'full_name': 'Trustworthy Foundation Models for Web Intelligence: Causal Perspectives and Challenges',
        'organizers': 'Haoang Chi (Tsinghua University), Qi Wang (Tsinghua University), Jiantong Jiang (University of Melbourne), Jiangchao Yao (Shanghai Jiaotong University), Feng Liu (University of Melbourne), Bo Han (Hong Kong Baptist University)',
        'url': 'https://www-tfm-causal.github.io/www2026-workshop/',
        'abstract': 'This workshop advances discussion on Trustworthy Foundation Models for the Web by introducing a causal perspective to improving the reliability, interpretability, and fairness of large-scale models. It convenes experts from machine learning, causal inference, web data mining, and social computing to establish a roadmap toward more robust, transparent, and ethically aligned Web AI systems.',
    },
    'TIME': {
        'full_name': '2nd International Workshop on Transformative Insights in Multi-faceted Evaluation',
        'organizers': 'Lei Wang (Griffith University & CSIRO), Md Zakir Hossain (Curtin University), Tom Gedeon (Curtin University), Syed Mohammed Shamsul Islam (ECU), Rafiqul Islam (Charles Sturt University), Yasmeen George (Monash University), Shreya Ghosh (University of Queensland)',
        'url': 'https://time.griffith.edu.au/workshop/time2026/',
        'abstract': 'TIME brings together domain experts to share insights on social network analysis, graph algorithms, web mining, semantics, security, privacy, fairness, and ethics on the web. The workshop invites survey, evaluation, or review papers that critically analyze models and datasets from diverse perspectives, complemented by invited talks from experts and industry leaders.',
    },
    'TML': {
        'full_name': 'International Workshop on Trustworthy Multimodal Learning for Social Media Analysis',
        'organizers': 'Jingwei Sun (ByteDance), Guosheng Lin (Nanyang Technological University), Fengmao Lv (Southwest Jiaotong University), Tao Liang (ByteDance), Junlin Fang (Southwest Jiaotong University)',
        'url': 'https://ttthhl.github.io/www2026-workshop/',
        'abstract': 'TML 2026 focuses on trustworthy multimodal learning methods for social media analysis, covering multimodal social media content analysis with LMMs, effective multimodal fusion and information alignment, and performance and safety evaluation of LMMs including quality of generated content, model hallucinations, and vulnerability to adversarial attacks.',
    },
    'WebAds': {
        'full_name': 'Emerging Trends in Web Advertising',
        'organizers': 'Ehsan Toreini (Samsung R&D Institute UK), Muadh Al Kalbani (Samsung R&D Institute UK)',
        'url': 'https://samsunginternet.github.io/webads26/',
        'abstract': 'The landscape of web advertising is undergoing a profound transformation, fueled by advancements in technologies that prioritize user privacy, AI-driven personalization, and immersive experiences. This workshop provides a platform for timely, responsible discussions among experts from advertising, privacy, data science, and related fields.',
    },
    'WebAndTheCity': {
        'full_name': '12th International Smart City Workshop - Data-Driven Smart Cities',
        'organizers': 'Leonidas Anthopoulos (University of Thessaly, Greece), Marijn Janssen (Delft University of Technology), Vishanth Weerakkody (University of Bradford, UK)',
        'url': 'https://webandthecity.home.blog/',
        'abstract': 'In the era of IoT, AI, and agentic AI integration, cities are being transformed into urban environments that use data as a foundational asset. This workshop explores how the Web supports smart city transformation and how technologies can improve urban decision-making, optimize services, and enhance citizen well-being.',
    },
    'WebST': {
        'full_name': 'The 2nd International Workshop on Spatio-Temporal Data Mining from the Web',
        'organizers': 'Yuxuan Liang (HKUST Guangzhou), Hao Xue (University of New South Wales), Ming Jin (Griffith University), Fei Wang (Institute of Computing Technology, CAS), Shirui Pan (Griffith University), Flora Salim (University of New South Wales)',
        'url': 'https://webst2026.netlify.app/',
        'abstract': 'A comprehensive workshop catering to professionals interested in sensing, mining, and understanding big and heterogeneous spatio-temporal data generated from the Web (social media posts, geotagged images, mobility traces) to tackle real-world challenges such as climate change, disaster response, urban planning, and location-based social networks.',
    },
    'ZABAPAD': {
        'full_name': 'Zero-knowledge Proof and Blockchain for Web 4.0: Advancing the Post-quantum and Decentralized Era',
        'organizers': 'Shiho Kim (Yonsei University), Roberto Di Pietro (KAUST), Davor Svetinovic (Khalifa University, UAE), KyungHi Chang (Inha University), Madhusudan Singh (Pennsylvania State University)',
        'url': 'https://zabapad.github.io',
        'abstract': 'ZABAPAD focuses on zero-knowledge technologies, blockchain infrastructure, and post-quantum readiness for the emerging Web 4.0 ecosystem. Topics include ZKP-based authentication, ZKML, Layer-2 proving/verification, TEE+ZK integration for verifiable compute, and post-quantum migration of identities, wallets, ledgers, and protocols across finance, mobility, healthcare, and AI/ML domains.',
    },
}

TUTORIAL_DETAILS = {
    'temporal information retrieval and question answering in the age of llms': {
        'presenters': 'Bhawna Piryani, Avishek Anand, and Adam Jatowt',
        'abstract': 'This tutorial provides a comprehensive overview of Temporal Information Retrieval (TIR) and Temporal Question Answering (TQA), addressing temporal relevance, reasoning, and adaptation in information access. It traces the evolution from early rule-based approaches to modern transformer and LLM architectures, highlighting how temporal modeling, reasoning, and retrieval-augmented generation are reshaping the field.',
    },
    'out-of-distribution generalized generative ai': {
        'presenters': 'Xin Wang, Yuwei Zhou, Zirui Pan, and Wenwu Zhu',
        'abstract': 'This tutorial disseminates recent research advancements in multi-modal generative AI, focusing on MLLMs and diffusion models. It covers solutions and future directions for challenges from shifting data distributions, emerging concepts, and evolving complex scenarios, including generalizable post-training techniques and unified multi-modal generation frameworks for dynamic open environments.',
    },
    'towards a responsible web: economic perspectives on fairness in information retrieval': {
        'presenters': 'Chen Xu, Clara Rus, Yuanna Liu, Marleen de Jonge, Jun Xu, and Maarten de Rijke',
        'abstract': 'Fairness is a crucial aspect of a responsible Web. This tutorial organizes fairness algorithms within an economic cube with dimensions: macro vs. micro, demand vs. supply, and short-term vs. long-term fairness. It draws parallels between IR systems and economic markets, demonstrating how IR fairness can be integrated into a structured economic framework with open problems and promising directions.',
    },
    'next-gen code development with collaborative ai agents': {
        'presenters': 'Shweta Garg, Behrooz Omidvar-Tehrani, Shengyu Fu, Gauthier Guinet, and Baishakhi Ray',
        'abstract': 'This tutorial explores AI-powered software development where LLMs function as collaborative agents that plan, code, test, and review alongside human developers. Using GitHub Copilot, Mistral Code and Kiro as exemplars, it covers multi-agent coordination, reflective collaboration, long-term memory, tool-integrated verification, and secure deployment patterns for modern engineering environments.',
    },
    'llm-enhanced web-centric spatio-temporal intelligence: methods, applications, and frontier research': {
        'presenters': 'Zijian Zhang, Hao Miao, Yuxuan Liang, Yan Zhao, and Irwin King',
        'abstract': 'This tutorial provides a comprehensive overview of LLM-Enhanced Web-Centric Spatio-Temporal Intelligence, organized at three levels: Location-level intelligence, Region-level intelligence, and broader spatio-temporal patterns. It presents methods, applications, and frontier research in the LLM era for Web-centric spatio-temporal data including geo-social media, LBS, and transportation records.',
    },
    'quantum-safe, efficient, and ai-enhanced blockchains for the web: a cooperative tutorial on quantum computing, blockchain applications, and data standards': {
        'presenters': 'Dongping Liu, Aoyu Zhang, and Luyao Zhang',
        'abstract': 'This tutorial explores how quantum computing and blockchain can jointly redefine trust, efficiency, and intelligence of next-generation Web systems. It covers principles of quantum computing and their implications for secure blockchain architectures, post-quantum cryptography, and culminates in a hands-on experience with cloud-based quantum computation through Amazon Braket.',
    },
    'conversational search: from fundamentals to frontiers in the age of agents': {
        'presenters': 'Chuan Meng, Fengran Mo, Mohammad Aliannejadi, Jeff Dalton, and Jian-Yun Nie',
        'abstract': 'This tutorial connects fundamentals with recent agentic paradigms in conversational search. It covers how LLMs enable multi-turn interactions to fulfill complex information needs, drive search systems toward agentic paradigms that can plan strategies, execute dynamic retrieval, and support autonomous behaviours. Designed for students, researchers, and practitioners from academia and industry.',
    },
    'bandits, llms, and agentic web': {
        'presenters': 'Djallel Bouneffouf, and Raphael Feraud',
        'abstract': 'This tutorial offers a comprehensive guide on using multiarmed bandit (MAB) algorithms to improve LLMs with a special focus on enabling agentic behavior. It covers foundational MAB concepts (epsilon-greedy, UCB, Thompson Sampling), integrating MAB with LLMs for text generation, and real-world case studies in content recommendation, dialogue generation, and personalized content creation.',
    },
    'unstructured to structured: building knowledge graphs from documents for web applications': {
        'presenters': 'Qiang Sun, Yihao Ding, Sirui Li and Wei Liu',
        'abstract': 'This tutorial presents methods for transforming unstructured Web content into structured Knowledge Graphs (KGs), covering information extraction across entities, relations, events, and spatio-temporal indices. It discusses hybrid systems combining LLMs with structured knowledge including LLM-driven KG construction, RAG over enterprise knowledge bases, and KG-augmented LLMs for grounded reasoning.',
    },
    'foundations for the agentic web: infrastructure, economics, and society': {
        'presenters': 'Ramesh Raskar, and Pradyumna Chari',
        'abstract': 'This tutorial provides a comprehensive framework for understanding the agentic web across three development phases: Foundations (discovery, identity, protocols), Agentic Economy (pricing, reputation, markets), and Agentic Society (population dynamics, governance, coordination). It draws on recent advances in registry architectures, protocol standards, and resolution mechanisms for agent ecosystems.',
    },
    'generalist model for structured data: foundations, frontiers and applications': {
        'presenters': 'Peng Cui, Xingxuan Zhang, Han-Jia Ye, Jintai Chen, and Shuyang Li',
        'abstract': 'Structured data is ubiquitous in web-scale and enterprise applications. This tutorial covers both conventional modeling paradigms and recent in-context learning (ICL)-based approaches for structured foundation models, discussing pretraining data generation, multi-task learning, and emerging directions including zero-shot inference and knowledge transfer across diverse structured settings.',
    },
    'robust graph learning on the web: challenges, methods, and applications': {
        'presenters': 'Xiang Ao, Yang Liu, Guansong Pang, Yuanhao Ding, Hezhe Qiao, Dawei Cheng, and Qing He',
        'abstract': 'This tutorial surveys strategies for robust graph learning on the Web, presenting a structured taxonomy of robustness threats (dynamic user behavior, incomplete content, adversarial interference, distribution shifts) and categorizing current approaches from data-level preprocessing to model-level adaptation. Includes real-world case studies from recommender systems to anomaly detection.',
    },
    'responsible prompting on the web: governance, mini-evaluation, and readiness with chatgpt': {
        'presenters': 'Manali Sharma and Ayush Garg',
        'abstract': 'This tutorial teaches a clear, repeatable workflow for responsible prompting in a browser-only setting with ChatGPT. It covers zero-shot vs. few-shot, chain-of-thought, role prompts, output formatting, multi-turn prompt chaining, and reverse prompting. Participants leave with prompt cards, a scoring rubric, and a deployment readiness one-pager documenting metrics, failure modes, and limitations.',
    },
}


def get_workshop_details(session_name):
    if not session_name:
        return None
    m = re.match(r'\(([^)]+)\)', session_name.strip())
    if m:
        code = m.group(1).strip()
        if code in WORKSHOP_DETAILS:
            return WORKSHOP_DETAILS[code]
        base = re.sub(r'\s*\d{4}$', '', code).rstrip('0123456789').rstrip("'").strip()
        if base in WORKSHOP_DETAILS:
            return WORKSHOP_DETAILS[base]
    return None


def get_tutorial_details(title):
    if not title:
        return None
    return TUTORIAL_DETAILS.get(title.strip().lower())


def should_skip_poster_block(block):
    return bool(re.search(r'\bYicheng Di\b', block or ''))


def normalize_title(text):
    if not text:
        return ''
    normalized = html_lib.unescape(str(text))
    normalized = unicodedata.normalize('NFKC', normalized)
    normalized = normalized.replace('\xa0', ' ')
    normalized = normalized.translate(str.maketrans({
        '\u2018': "'",
        '\u2019': "'",
        '\u201c': '"',
        '\u201d': '"',
        '\u2013': '-',
        '\u2014': '-',
        '\u2011': '-',
        '\u2212': '-',
    }))
    normalized = re.sub(r'[\u200b\u200c\u200d\u2060\ufeff]', '', normalized)
    normalized = re.sub(r'\s+', ' ', normalized).strip().lower()
    return normalized


def build_title_url_map(workbook):
    title_url_map = {}
    for sheet_name in ('Mapping to URL Main Proceedings', 'Mapping to URL Comp Proceedings'):
        if sheet_name not in workbook.sheetnames:
            continue
        ws_map = workbook[sheet_name]
        for row in ws_map.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            title = row[1]
            url = row[2]
            if not title or not url:
                continue
            normalized_title = normalize_title(title)
            if normalized_title and normalized_title not in title_url_map:
                title_url_map[normalized_title] = str(url).strip()
    return title_url_map


MISSING_POSTER_PAPERS = {
    'ShortPapers5': [{
        'title': 'Linguistic Signatures for Enhanced Emotion Detection',
        'authors': 'Florian Lecourt, Madalina Croitoru and Konstantin Todorov',
        'acm_url': None,
    }],
}


# ─────────────────────────────────────────────
# 2. Parse Excel
# ─────────────────────────────────────────────

wb = openpyxl.load_workbook(
    'program/Detailed WebConf 2026 Program.xlsx', data_only=True
)
ws = wb['Program']
TITLE_URL_MAP = build_title_url_map(wb)

all_rows = [list(row) for row in ws.iter_rows(values_only=True)]


def normalize_hall_label(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    hall_num = re.search(r'\d+', text)
    if hall_num:
        return f'Hall {hall_num.group(0)}'
    return re.sub(r'\bRooms?\b', 'Hall', text, flags=re.IGNORECASE)


HALL_LABELS = {}
for col_idx in range(1, min(27, len(all_rows[1]))):
    hall_label = normalize_hall_label(all_rows[1][col_idx])
    if hall_label:
        HALL_LABELS[col_idx] = hall_label

last_row = 0
for i, row in enumerate(all_rows):
    if any(v is not None for v in row):
        last_row = i


def is_special_event(slot_code):
    specials = ['break', 'keynote', 'conference opening', 'town hall', 'plenary']
    if slot_code is None:
        return False
    return any(slot_code.lower().startswith(s) for s in specials)


def is_poster_session(slot_code):
    if slot_code is None:
        return False
    return str(slot_code).lower().startswith('posters and demos')


def track_label(code):
    if code is None:
        return ''
    c = str(code).strip()
    mappings = {
        'tut': 'Tutorial', 'wk': 'Workshop',
        'Econ': 'Economics', 'Graph': 'Graph Neural Networks',
        'Search': 'Search & Retrieval', 'Semantics': 'Semantic Web',
        'Social': 'Social Web', 'Systems': 'Web Systems',
        'Security': 'Security & Privacy', 'Mining': 'Data Mining',
        'RecSys': 'Recommender Systems', 'Industry': 'Industry Track',
        'History': 'History of the Web', 'Web4Good': 'Web4Good',
        'RespWeb': 'Responsible Web', 'PhD Symposium': 'PhD Symposium',
        'Web4All': 'Web4All',
        'Demos': 'Demos', 'ShortPapers': 'Short Papers',
        'EconPoster': 'Economics Posters', 'GraphPoster': 'Graph Posters',
        'SearchPoster': 'Search Posters', 'SemanticsPoster': 'Semantics Posters',
        'SecurityPoster': 'Security Posters', 'SocialPoster': 'Social Posters',
        'SystemsPoster': 'Systems Posters', 'RecSysPoster': 'RecSys Posters',
        'MiningPoster': 'Mining Posters', 'IndustryPoster': 'Industry Posters',
        'Web4GoodPoster': 'Web4Good Posters', 'RespWebPoster': 'RespWeb Posters',
        'BestPaper': 'Best Paper Candidates',
    }
    for prefix, label in mappings.items():
        if c.startswith(prefix):
            return label
    return c


def track_color_class(code):
    if code is None:
        return 'track-other'
    c = str(code).strip()
    if c.startswith('tut'):       return 'track-tutorial'
    if c.startswith('wk'):        return 'track-workshop'
    if c.startswith('Econ'):      return 'track-econ'
    if c.startswith('Graph'):     return 'track-graph'
    if c.startswith('Search'):    return 'track-search'
    if c.startswith('Semantic'):  return 'track-semantics'
    if c.startswith('Social'):    return 'track-social'
    if c.startswith('System'):    return 'track-systems'
    if c.startswith('Security'):  return 'track-security'
    if c.startswith('Mining'):    return 'track-mining'
    if c.startswith('RecSys'):    return 'track-recsys'
    if c.startswith('Industry'):  return 'track-industry'
    if c.startswith('History'):   return 'track-history'
    if c.startswith('Web4Good'):  return 'track-web4good'
    if c.startswith('RespWeb'):   return 'track-respweb'
    if c.startswith('PhD'):       return 'track-phd'
    if c.startswith('Web4All'):   return 'track-web4all'
    if c.startswith('Demos'):     return 'track-demos'
    if c.startswith('ShortP'):    return 'track-short'
    if c.startswith('BestP'):     return 'track-best'
    return 'track-other'


def workshop_listing_header(title, session_name):
    normalized_title = normalize_title(title)
    details = get_workshop_details(session_name)
    if details:
        full_name = details.get('full_name')
        if full_name and normalize_title(full_name) in normalized_title:
            return True
    if session_name:
        code_match = re.match(r'\(([^)]+)\)', session_name.strip())
        if code_match and normalize_title(code_match.group(1)) in normalized_title:
            return True
    return False


def attach_paper_urls(program):
    for day in program:
        for slot in day.get('slots', []):
            for sess in slot.get('sessions') or []:
                for paper_key in ('papers', 'proceedings_papers'):
                    for paper in sess.get(paper_key, []):
                        normalized_title = normalize_title(paper.get('title'))
                        if normalized_title in TITLE_URL_MAP:
                            paper['acm_url'] = TITLE_URL_MAP[normalized_title]


program = []
current_day = None
i = 0

while i <= last_row:
    row = all_rows[i]
    col_a = row[0]
    col_b = row[1] if len(row) > 1 else None

    if i < 2:
        i += 1
        continue

    if isinstance(col_a, datetime):
        current_day = {'date': col_a, 'slots': []}
        program.append(current_day)
        i += 1
        continue

    if col_a and isinstance(col_a, str) and current_day is not None:
        time_str = col_a.strip()

        codes = []
        for j in range(1, 27):
            val = row[j] if j < len(row) else None
            if val is not None:
                codes.append({'col_idx': j, 'code': str(val).strip()})

        only_b = (len(codes) == 1 and codes[0]['col_idx'] == 1)
        special = is_special_event(col_b) if only_b else False
        poster = is_poster_session(col_b) if only_b else False

        slot = {
            'time': time_str,
            'special': special,
            'poster': poster,
            'event': col_b if (special or poster) else None,
            'sessions': [] if not (special or poster) else None,
            'detail_rows': []
        }

        if not special and not poster:
            for c in codes:
                slot['sessions'].append({
                    'col_idx': c['col_idx'],
                    'code': c['code'],
                    'hall': HALL_LABELS.get(c['col_idx']),
                    'name': None,
                    'papers': [],
                    'proceedings_papers': [],
                    'url': None,
                })

        current_day['slots'].append(slot)

        i += 1
        detail_rows = []
        while i <= last_row:
            nr = all_rows[i]
            if nr[0] is not None:
                break
            dr = {}
            for j in range(1, 27):
                v = nr[j] if j < len(nr) else None
                if v is not None and isinstance(v, str):
                    dr[j] = str(v)
            if dr:
                detail_rows.append(dr)
            i += 1

        slot['detail_rows'] = detail_rows

        if poster and len(detail_rows) >= 2:
            dr0 = detail_rows[0]
            dr1 = detail_rows[1]
            sub_sessions = []
            for j in sorted(dr0.keys()):
                code_val = dr0[j].strip()
                if not code_val:
                    continue
                papers = []
                if j in dr1:
                    blocks = re.split(r'\n\s*\n', dr1[j].strip())
                    for block in blocks:
                        block = block.strip()
                        if not block:
                            continue
                        if should_skip_poster_block(block):
                            continue
                        blines = block.split('\n')
                        title_p = blines[0].strip()
                        if re.match(r'WITHDRAWN', title_p, re.IGNORECASE):
                            continue
                        authors_p = ' '.join(l.strip() for l in blines[1:] if l.strip())
                        papers.append({
                            'title': title_p,
                            'authors': authors_p,
                            'acm_url': None
                        })
                for missing_paper in reversed(MISSING_POSTER_PAPERS.get(code_val, [])):
                    if not any(p['title'] == missing_paper['title'] for p in papers):
                        papers.insert(0, missing_paper)
                sub_sessions.append({
                    'col_idx': j,
                    'code': code_val,
                    'name': track_label(code_val),
                    'papers': papers,
                    'proceedings_papers': [],
                    'url': None,
                })
            slot['sessions'] = sub_sessions

        elif not special and not poster and slot['sessions']:
            sessions_by_col = {s['col_idx']: s for s in slot['sessions']}
            n_detail = len(detail_rows)

            if n_detail == 1:
                dr0 = detail_rows[0]
                for j, text in dr0.items():
                    if j in sessions_by_col:
                        if '\n\n' in text.strip():
                            blocks = re.split(r'\n\s*\n', text.strip())
                            sessions_by_col[j]['name'] = None
                            for block in blocks:
                                block = block.strip()
                                if not block:
                                    continue
                                blines = block.split('\n')
                                title = blines[0].strip()
                                authors = ' '.join(l.strip() for l in blines[1:] if l.strip())
                                sessions_by_col[j]['papers'].append({
                                    'title': title, 'authors': authors, 'acm_url': None
                                })
                        else:
                            lines = text.split('\n')
                            first = lines[0].strip()
                            if re.match(r'https?://', first) and len(lines) == 1:
                                sessions_by_col[j]['url'] = first
                            else:
                                sessions_by_col[j]['name'] = first
                                rest = '\n'.join(lines[1:]).strip()
                                urls = re.findall(r'https?://\S+', rest)
                                if urls:
                                    sessions_by_col[j]['url'] = urls[0]
                                non_url = re.sub(r'https?://\S+', '', rest).strip()
                                sessions_by_col[j]['papers'].append({
                                    'title': first, 'authors': non_url, 'acm_url': None
                                })

            elif n_detail >= 2:
                dr0 = detail_rows[0]
                for j, text in dr0.items():
                    if j in sessions_by_col:
                        sessions_by_col[j]['name'] = text.strip()

                dr1 = detail_rows[1]
                for j, text in dr1.items():
                    if j in sessions_by_col:
                        session = sessions_by_col[j]
                        is_workshop = session['code'].startswith('wk')
                        blocks = re.split(r'\n\s*\n', text.strip())
                        for index, block in enumerate(blocks):
                            block = block.strip()
                            if not block:
                                continue
                            blines = block.split('\n')
                            title = blines[0].strip()
                            if is_workshop and index == 0 and workshop_listing_header(title, session.get('name')):
                                continue
                            authors = ' '.join(l.strip() for l in blines[1:] if l.strip())
                            target_key = 'proceedings_papers' if is_workshop else 'papers'
                            session[target_key].append({
                                'title': title, 'authors': authors, 'acm_url': None
                            })
        continue

    i += 1


attach_paper_urls(program)


# ─────────────────────────────────────────────
# 3. HTML helpers
# ─────────────────────────────────────────────

def esc(s):
    return html_lib.escape(str(s)) if s else ''


def render_paper_link(paper):
    acm_url = paper.get('acm_url')
    href = esc(acm_url) if acm_url else '#'
    target_attrs = ' target="_blank" rel="noopener noreferrer"' if acm_url else ''
    data_acm = esc(acm_url) if acm_url else ''
    return f'<a href="{href}" class="paper-title acm-link-placeholder" data-acm="{data_acm}"{target_attrs}>{esc(paper["title"])}</a>'


def render_paper_list_html(papers):
    items = []
    for paper in papers:
        title_link = render_paper_link(paper)
        if paper.get('authors'):
            items.append(f'<li class="paper-item"><span class="paper-title-wrap">{title_link}</span><span class="paper-authors">{esc(paper["authors"])}</span></li>')
        else:
            items.append(f'<li class="paper-item"><span class="paper-title-wrap">{title_link}</span></li>')
    return '<ul class="paper-list">' + ''.join(items) + '</ul>'


def render_hall_label(sess):
    hall = sess.get('hall')
    if not hall:
        return ''
    code = str(sess.get('code') or '')
    if code.startswith('Demo') or 'Poster' in code:
        return ''
    return f'<div class="session-hall">{esc(str(hall).upper())}</div>'


def day_id(day):
    return day['date'].strftime('day-%Y-%m-%d')

def format_day_tab(day):
    d = day['date']
    return f"{d.strftime('%A')}<br><span>{d.strftime('%-d %b')}</span>"

def format_day_heading(day):
    return day['date'].strftime('%A, %B %-d, %Y')

SPECIAL_EVENT_CLASS = {
    'break': 'slot-break',
    'keynote': 'slot-keynote',
    'conference opening': 'slot-opening',
    'town hall': 'slot-closing',
    'plenary': 'slot-plenary',
    'posters and demos': 'slot-posters',
}

def special_class(event):
    if event is None:
        return 'slot-special'
    el = event.lower()
    for k, v in SPECIAL_EVENT_CLASS.items():
        if el.startswith(k):
            return v
    return 'slot-special'

def item_label(code, count):
    c = str(code)
    singular, plural = (
        ('poster',       'posters')      if 'Poster' in c else
        ('demo',         'demos')        if c.startswith('Demo') else
        ('short paper',  'short papers') if c.startswith('ShortP') else
        ('paper',        'papers')       if c.startswith('PhD') else
        ('item',         'items')        if c.startswith('History') or c.startswith('Web4All') else
        ('tutorial',     'tutorials')    if c.startswith('tut') else
        ('workshop',     'workshops')    if c.startswith('wk') else
        ('paper',        'papers')
    )
    return f'{count} {singular if count == 1 else plural}'


def parse_slot_datetimes(day, time_str):
    matches = re.findall(r'(\d{1,2}):(\d{2})', str(time_str or ''))
    if len(matches) < 2:
        start = day['date'].replace(hour=9, minute=0, second=0, microsecond=0)
        return start, start + timedelta(hours=1)
    (sh, sm), (eh, em) = matches[0], matches[1]
    start = day['date'].replace(hour=int(sh), minute=int(sm), second=0, microsecond=0)
    end = day['date'].replace(hour=int(eh), minute=int(em), second=0, microsecond=0)
    if end <= start:
        end += timedelta(days=1)
    return start, end


def ics_escape(value):
    return str(value or '').replace('\\', '\\\\').replace(';', '\\;').replace(',', '\\,').replace('\n', '\\n')


def safe_filename(value):
    slug = re.sub(r'[^A-Za-z0-9]+', '-', str(value or 'session')).strip('-').lower()
    return (slug[:70] or 'session') + '.ics'


def calendar_link(day, time_str, title, description='', slot_idx=0, item_idx=0):
    start, end = parse_slot_datetimes(day, time_str)
    uid_seed = f'www2026-{day["date"].strftime("%Y%m%d")}-{slot_idx}-{item_idx}'
    lines = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//The Web Conference 2026//Program//EN',
        'CALSCALE:GREGORIAN',
        'METHOD:PUBLISH',
        'BEGIN:VEVENT',
        f'UID:{uid_seed}@thewebconf.org',
        'DTSTAMP:20260527T000000Z',
        f'DTSTART;TZID=Asia/Dubai:{start.strftime("%Y%m%dT%H%M%S")}',
        f'DTEND;TZID=Asia/Dubai:{end.strftime("%Y%m%dT%H%M%S")}',
        f'SUMMARY:{ics_escape(title)}',
        f'DESCRIPTION:{ics_escape(description)}',
        'LOCATION:Dubai World Trade Centre, Dubai, UAE',
        'END:VEVENT',
        'END:VCALENDAR',
    ]
    href = 'data:text/calendar;charset=utf-8,' + quote('\r\n'.join(lines))
    filename = safe_filename(f'{title}-{start.strftime("%Y%m%d-%H%M")}-{item_idx}')
    return (
        f'<a class="calendar-link" href="{href}" download="{esc(filename)}">'
        f'<i class="far fa-calendar-plus"></i><span>Calendar</span></a>'
    )


def paper_details_text(papers):
    lines = []
    for p in papers:
        title = p.get('title', '')
        authors = p.get('authors', '')
        if title and authors:
            lines.append(f'- {title} - {authors}')
        elif title:
            lines.append(f'- {title}')
    return '\n'.join(lines)


def calendar_items_heading(code):
    code = str(code)
    if 'Poster' in code:
        return 'Posters'
    if code.startswith('Demo'):
        return 'Demos'
    if code.startswith('ShortP'):
        return 'Short papers'
    return 'Papers/items'


def session_calendar_description(code, tlabel, name, papers, details=None, url=''):
    parts = [f'The Web Conference 2026: {tlabel}.']
    if details:
        if details.get('full_name') and details.get('full_name') != name:
            parts.append(f'\nTitle: {details["full_name"]}')
        if details.get('organizers'):
            parts.append(f'\nOrganizers: {details["organizers"]}')
        if details.get('presenters'):
            parts.append(f'\nPresenters: {details["presenters"]}')
        detail_url = details.get('url') or url
        if detail_url:
            parts.append(f'\nWebsite: {detail_url}')
        if details.get('abstract'):
            parts.append(f'\nAbstract: {details["abstract"]}')
    elif papers:
        parts.append(f'\n{calendar_items_heading(code)}:\n{paper_details_text(papers)}')
    elif url:
        parts.append(f'\nWebsite: {url}')
    return ''.join(parts)


def render_session_card_overview(sess):
    code = sess['code']
    name = sess['name'] or track_label(code)
    color_cls = track_color_class(code)
    tlabel = track_label(code)
    url = sess.get('url')
    hall_html = render_hall_label(sess)
    title_html = esc(name)
    if url and not code.startswith('tut') and not code.startswith('wk'):
        title_html = f'<a href="{esc(url)}" target="_blank" class="session-ext-link">{esc(name)} <i class="fas fa-external-link-alt"></i></a>'
    return f'''<div class="session-card {color_cls}">
  <div class="session-track-badge">{esc(tlabel)}</div>
  {hall_html}
  <div class="session-name">{title_html}</div>
</div>'''


def render_session_card_full(sess, day, slot, slot_idx, sess_idx):
    code = sess['code']
    name = sess['name'] or track_label(code)
    color_cls = track_color_class(code)
    tlabel = track_label(code)
    papers = sess['papers']
    proceedings_papers = sess.get('proceedings_papers') or []
    url = sess.get('url')
    hall_html = render_hall_label(sess)
    collapse_id = f"collapse-{slot_idx}-{sess_idx}"
    proceedings_collapse_id = f"collapse-acm-{slot_idx}-{sess_idx}"

    is_tutorial = code.startswith('tut')
    is_workshop = code.startswith('wk')

    title_html = esc(name)
    if url and not is_tutorial and not is_workshop:
        title_html = f'<a href="{esc(url)}" target="_blank" class="session-ext-link">{esc(name)} <i class="fas fa-external-link-alt"></i></a>'

    papers_section = ''
    header_extra = ''
    calendar_details = None

    if is_tutorial or is_workshop:
        details = get_workshop_details(name) if is_workshop else get_tutorial_details(name)
        calendar_details = details
        detail_content = ''
        if details:
            if is_workshop:
                ws_url = details.get('url', url or '')
                detail_content += f'<p class="detail-full-name"><strong>{esc(details["full_name"])}</strong></p>'
                if details.get('organizers'):
                    detail_content += f'<p class="detail-organizers"><i class="fas fa-users"></i> {esc(details["organizers"])}</p>'
                if ws_url:
                    detail_content += f'<p class="detail-link"><i class="fas fa-globe"></i> <a href="{esc(ws_url)}" target="_blank">{esc(ws_url)}</a></p>'
                if details.get('abstract'):
                    detail_content += f'<p class="detail-abstract"><strong>About:</strong> {esc(details["abstract"])}</p>'
            else:
                detail_content += f'<p class="detail-full-name"><strong>{esc(name)}</strong></p>'
                if details.get('presenters'):
                    detail_content += f'<p class="detail-organizers"><i class="fas fa-chalkboard-teacher"></i> <strong>Presenters:</strong> {esc(details["presenters"])}</p>'
                if details.get('abstract'):
                    detail_content += f'<p class="detail-abstract"><strong>About:</strong> {esc(details["abstract"])}</p>'
        else:
            detail_content += f'<p class="detail-full-name"><strong>{esc(name)}</strong></p>'
            calendar_details = {'full_name': name}
            if papers and papers[0].get('authors'):
                label = 'Presenters' if is_tutorial else 'Organizers'
                detail_content += f'<p class="detail-organizers"><strong>{label}:</strong> {esc(papers[0]["authors"])}</p>'
                calendar_details[label.lower()] = papers[0]['authors']
            if url:
                detail_content += f'<p class="detail-link"><i class="fas fa-globe"></i> <a href="{esc(url)}" target="_blank">{esc(url)}</a></p>'
                calendar_details['url'] = url

        toggle = f' data-bs-toggle="collapse" data-bs-target="#{collapse_id}" aria-expanded="false" aria-controls="{collapse_id}"'
        papers_section = f'<div class="collapse" id="{collapse_id}"><div class="papers-container">{detail_content}</div></div>'
        header_extra = f'<button class="papers-toggle"{toggle}><span class="papers-count">Details</span><i class="fas fa-chevron-down toggle-icon"></i></button>'
        if is_workshop and proceedings_papers:
            proceedings_toggle = f' data-bs-toggle="collapse" data-bs-target="#{proceedings_collapse_id}" aria-expanded="false" aria-controls="{proceedings_collapse_id}"'
            proceedings_html = render_paper_list_html(proceedings_papers)
            papers_section += f'<div class="collapse" id="{proceedings_collapse_id}"><div class="papers-container">{proceedings_html}</div></div>'
            header_extra += f'<button class="papers-toggle"{proceedings_toggle}><span class="papers-count">Papers in ACM DL</span><i class="fas fa-chevron-down toggle-icon"></i></button>'

    elif papers:
        papers_html = render_paper_list_html(papers)
        toggle = f' data-bs-toggle="collapse" data-bs-target="#{collapse_id}" aria-expanded="false" aria-controls="{collapse_id}"'
        paper_label = item_label(code, len(papers))
        papers_section = f'<div class="collapse" id="{collapse_id}"><div class="papers-container">{papers_html}</div></div>'
        header_extra = f'<button class="papers-toggle"{toggle}><span class="papers-count">{paper_label}</span><i class="fas fa-chevron-down toggle-icon"></i></button>'

    header_extra += calendar_link(
        day,
        slot['time'],
        name,
        session_calendar_description(code, tlabel, name, papers, calendar_details, url),
        slot_idx,
        sess_idx,
    )
    actions_html = f'<div class="session-actions">{header_extra}</div>'

    return f'''<div class="session-card {color_cls}">
  <div class="session-card-header">
    <div class="session-track-badge">{esc(tlabel)}</div>
    {hall_html}
    <div class="session-name">{title_html}</div>
    {actions_html}
  </div>
  {papers_section}
</div>'''


def render_keynote_block_overview(event, slot_idx):
    sc = special_class(event)
    kd = keynote_lookup(event)
    if kd:
        return (f'<div class="slot-special-block {sc}">'
                f'<div class="keynote-speaker-name">{esc(event)}</div>'
                f'<div class="keynote-talk-title">{esc(kd["title"])}</div>'
                f'</div>')
    return f'<div class="slot-special-block {sc}">{esc(event)}</div>'


def render_keynote_block_full(event, day, slot, slot_idx):
    sc = special_class(event)
    kd = keynote_lookup(event)
    if not kd:
        return f'<div class="slot-special-block {sc}">{esc(event)}</div>'
    collapse_id = f"keynote-detail-{slot_idx}"
    detail_html = (
        f'<p class="keynote-detail-title"><strong>Talk Title:</strong> {esc(kd["title"])}</p>'
        f'<p class="keynote-detail-abstract"><strong>Abstract:</strong> {esc(kd["abstract"])}</p>'
        f'<p class="keynote-detail-bio"><strong>Bio:</strong> {esc(kd["bio"])}</p>'
    )
    cal_description = (
        f'The Web Conference 2026 keynote.\n'
        f'Talk Title: {kd["title"]}\n'
        f'Abstract: {kd["abstract"]}\n'
        f'Bio: {kd["bio"]}'
    )
    cal_link = calendar_link(day, slot['time'], event, cal_description, slot_idx, 0)
    action_html = (
        f'<div class="event-actions">'
        f'<button class="papers-toggle keynote-details-btn" data-bs-toggle="collapse" data-bs-target="#{collapse_id}" aria-expanded="false" aria-controls="{collapse_id}">'
        f'<span class="papers-count">Details</span><i class="fas fa-chevron-down toggle-icon"></i>'
        f'</button>'
        f'{cal_link}'
        f'</div>'
    )
    return (
        f'<div class="slot-special-block {sc}">'
        f'<div class="keynote-header-row">'
        f'<div>'
        f'<div class="keynote-speaker-name">{esc(event)}</div>'
        f'<div class="keynote-talk-title">{esc(kd["title"])}</div>'
        f'</div>'
        f'{action_html}'
        f'</div>'
        f'<div class="collapse" id="{collapse_id}"><div class="keynote-expand-content">{detail_html}</div></div>'
        f'</div>'
    )


def render_special_block_full(event, day, slot, slot_idx):
    if special_class(event) == 'slot-break':
        return f'<div class="slot-special-block {special_class(event)}">{esc(event)}</div>'
    return (
        f'<div class="slot-special-block {special_class(event)}">'
        f'<div class="special-event-row"><span>{esc(event)}</span>{calendar_link(day, slot["time"], event, "The Web Conference 2026", slot_idx, 0)}</div>'
        f'</div>'
    )


def render_poster_session(slot, is_full_schedule, slot_idx, day=None):
    event_name = slot['event']
    sc = special_class(event_name)
    sub_sessions = slot.get('sessions') or []
    header_content = esc(event_name)
    if is_full_schedule and day is not None:
        header_content = (
            f'<div class="special-event-row"><span>{esc(event_name)}</span>'
            f'{calendar_link(day, slot["time"], event_name, "The Web Conference 2026 poster and demo session", slot_idx, 0)}</div>'
        )
    header_block = f'<div class="slot-special-block {sc} poster-session-header">{header_content}</div>'
    if not sub_sessions:
        return header_block
    cards = ''
    for si, sess in enumerate(sub_sessions):
        if is_full_schedule:
            cards += render_session_card_full(sess, day, slot, slot_idx, si + 1) + '\n'
        else:
            cards += render_session_card_overview(sess) + '\n'
    return f'{header_block}\n<div class="session-grid poster-sub-grid">{cards}</div>'


# ─────────────────────────────────────────────
# 4. CSS / scripts
# ─────────────────────────────────────────────

COMMON_HEAD_LINKS = '''
    <link rel="stylesheet" href="../css/styles.css">
    <link rel="stylesheet" href="../css/tab-menu.css">
    <link rel="stylesheet" href="../css/dropdown-menu.css">
    <link rel="stylesheet" href="../css/header-layout.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css">
'''

PROGRAM_CSS = '''<style>
.program-page { padding: 30px 0 60px; }
.day-tabs {
  display: grid; grid-template-columns: repeat(5, minmax(0, 1fr)); gap: 6px;
  margin-bottom: 30px; border-bottom: 2px solid #e0e0e0; padding-bottom: 0;
  background: white;
}
.day-tab-btn {
  background: #f4f4f4; border: none; border-radius: 6px 6px 0 0;
  padding: 10px 18px; cursor: pointer; font-size: 0.9rem; font-weight: 600;
  color: #555; transition: all .2s; text-align: center; line-height: 1.3;
  border-bottom: 3px solid transparent; margin-bottom: -2px;
}
.day-tab-btn:hover { background: #ede5ff; color: #6a00ff; }
.day-tab-btn.active { background: white; color: #6a00ff; border-bottom: 3px solid #6a00ff; }
.day-tab-btn:focus-visible,
.papers-toggle:focus-visible,
.calendar-link:focus-visible {
  outline: 3px solid rgba(106,0,255,.22); outline-offset: 2px;
}
.day-tab-btn span { font-size: 0.8rem; font-weight: 400; display: block; }
.day-panel { display: none; }
.day-panel.active { display: block; }
.time-slot { margin-bottom: 20px; }
.time-label {
  display: inline-flex; align-items: center; gap: 6px;
  color: #555; font-size: 0.82rem; font-weight: 700;
  letter-spacing: .4px; margin-bottom: 10px; padding: 4px 0;
}
.time-label i { opacity: .5; font-size: 0.78rem; }
.slot-special-block {
  border-radius: 6px; padding: 13px 18px; margin-bottom: 4px;
  font-size: 0.95rem; font-weight: 600; line-height: 1.4; color: #1a1a1a;
  background: #fff; border: 1px solid #e8e8e8; border-left: 4px solid #bbb;
  box-shadow: 0 1px 4px rgba(0,0,0,.05);
}
.slot-keynote  { border-left-color: #6a00ff; }
.slot-break    { border-left-color: #bbb; font-weight: 400; font-size: 0.85rem; }
.slot-opening  { border-left-color: #00a855; }
.slot-closing  { border-left-color: #00a855; }
.slot-plenary  { border-left-color: #e08000; }
.slot-posters  { border-left-color: #0078d4; }
/* Keynote expanded */
.keynote-header-row {
  display: flex; align-items: flex-start; justify-content: space-between; gap: 12px;
}
.keynote-header-row > div:first-child { flex: 1; min-width: 0; }
.special-event-row {
  display: flex; align-items: center; justify-content: space-between; gap: 12px;
}
.event-actions,
.session-actions {
  display: flex; flex-wrap: wrap; align-items: center; gap: 6px;
}
.event-actions { justify-content: flex-end; flex-shrink: 0; }
.keynote-speaker-name { font-size: 0.95rem; font-weight: 700; color: #1a1a1a; }
.keynote-talk-title   { font-size: 0.88rem; font-weight: 500; color: #555; margin-top: 3px; font-style: italic; }
.keynote-details-btn  { flex-shrink: 0; }
.calendar-link {
  display: inline-flex; align-items: center; justify-content: center; gap: 5px;
  height: 28px; box-sizing: border-box; align-self: flex-start; flex-shrink: 0;
  border: 1px solid #d9d9d9; border-radius: 5px; padding: 0 8px;
  background: #fff; color: #444; font-size: 0.74rem; font-weight: 600;
  line-height: 1; text-decoration: none; transition: background .2s, border-color .2s, color .2s;
}
.calendar-link:hover { background: #f5f0ff; border-color: #c8b5ff; color: #5c22d6; text-decoration: none; }
.calendar-link i { font-size: 0.78rem; }
.keynote-expand-content {
  margin-top: 12px; padding-top: 12px; border-top: 1px solid #e8e0ff;
  font-size: 0.84rem; color: #333; line-height: 1.6; font-weight: 400;
}
.keynote-detail-title { margin: 0 0 8px; font-style: italic; color: #555; }
.keynote-detail-abstract { margin: 8px 0; }
.keynote-detail-bio { margin: 8px 0 0; }
.keynote-expand-content strong { color: #1a1a1a; }
/* Poster sessions */
.poster-session-header { margin-bottom: 10px; background: #f8fbff; }
.poster-sub-grid { margin-top: 8px; }
/* Workshop/tutorial detail panel */
.detail-full-name { margin: 0 0 6px; }
.detail-organizers { font-size: 0.8rem; color: #555; margin: 4px 0; }
.detail-organizers i { margin-right: 5px; color: #888; }
.detail-link { font-size: 0.8rem; margin: 4px 0; }
.detail-link i { margin-right: 5px; color: #888; }
.detail-link a { color: #1565c0; word-break: break-all; }
.detail-link a:hover { text-decoration: underline; }
.detail-abstract { font-size: 0.81rem; color: #444; margin: 6px 0 0; line-height: 1.55; }
.detail-abstract strong { color: #1a1a1a; }
/* Session grid */
.session-grid {
  display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px;
}
@media (max-width: 768px) {
  .session-grid { grid-template-columns: 1fr 1fr; }
  .day-tabs { grid-template-columns: repeat(2, minmax(0, 1fr)); }
  .day-tab-btn { padding: 8px 12px; font-size: 0.82rem; }
  .slot-special-block { font-size: 0.88rem; padding: 11px 14px; }
  .day-heading { font-size: 1.1rem; }
  .keynote-header-row { flex-direction: column; gap: 8px; }
  .event-actions { justify-content: flex-start; }
  .special-event-row { align-items: flex-start; }
}
@media (max-width: 480px) {
  .session-grid { grid-template-columns: 1fr; }
  .day-tabs { grid-template-columns: 1fr; gap: 4px; }
  .day-tab-btn { padding: 7px 9px; font-size: 0.78rem; }
  .time-label { font-size: 0.78rem; }
  .program-page { padding: 16px 0 40px; }
}
/* Session cards */
.session-card {
  border-radius: 8px; padding: 12px 14px;
  background: white; border: 1px solid #e0e0e0;
  box-shadow: 0 1px 4px rgba(0,0,0,.06);
  transition: border-color .2s, box-shadow .2s, transform .1s;
  display: flex; flex-direction: column; gap: 6px;
}
.session-card:hover { border-color: #d2c4ff; box-shadow: 0 4px 14px rgba(0,0,0,.13); transform: translateY(-1px); }
.session-card-header { display: flex; flex-direction: column; gap: 6px; }
.session-track-badge {
  font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
  letter-spacing: .7px; padding: 2px 8px; border-radius: 20px;
  display: inline-block; align-self: flex-start;
}
.session-name { font-size: 0.88rem; font-weight: 600; color: #1a1a1a; line-height: 1.4; }
.session-hall {
  font-size: 1.02rem; font-weight: 900; text-transform: uppercase; letter-spacing: .7px;
  color: #111; line-height: 1.1;
}
.session-ext-link { color: inherit; text-decoration: none; }
.session-ext-link:hover { text-decoration: underline; }
.session-ext-link .fas { font-size: 0.7rem; opacity: 0.7; }
/* Track colors */
.track-tutorial .session-track-badge  { background:#f3e8ff; color:#6a00ff; }
.track-workshop .session-track-badge  { background:#e8eaff; color:#2d3ad0; }
.track-econ     .session-track-badge  { background:#fff3e0; color:#b06000; }
.track-graph    .session-track-badge  { background:#e3f2fd; color:#0d47a1; }
.track-search   .session-track-badge  { background:#e0f7fa; color:#006064; }
.track-semantics .session-track-badge { background:#e8f5e9; color:#1b5e20; }
.track-social   .session-track-badge  { background:#fce4ec; color:#880e4f; }
.track-systems  .session-track-badge  { background:#ede7f6; color:#311b92; }
.track-security .session-track-badge  { background:#ffebee; color:#b71c1c; }
.track-mining   .session-track-badge  { background:#e8f5e9; color:#33691e; }
.track-recsys   .session-track-badge  { background:#e8eaf6; color:#1a237e; }
.track-industry .session-track-badge  { background:#fafafa; color:#424242; border:1px solid #bdbdbd; }
.track-history  .session-track-badge  { background:#fff8e1; color:#6d4c41; }
.track-web4good .session-track-badge  { background:#e0f2f1; color:#004d40; }
.track-respweb  .session-track-badge  { background:#fbe9e7; color:#bf360c; }
.track-phd      .session-track-badge  { background:#f9fbe7; color:#558b2f; }
.track-web4all  .session-track-badge  { background:#e1f5fe; color:#01579b; }
.track-demos    .session-track-badge  { background:#e8eaf6; color:#283593; }
.track-short    .session-track-badge  { background:#f3e5f5; color:#6a1b9a; }
.track-best     .session-track-badge  { background:#fff9c4; color:#827717; }
.track-other    .session-track-badge  { background:#f5f5f5; color:#616161; }
.track-tutorial  { border-left:3px solid #6a00ff; }
.track-workshop  { border-left:3px solid #2d3ad0; }
.track-econ      { border-left:3px solid #f0a000; }
.track-graph     { border-left:3px solid #1565c0; }
.track-search    { border-left:3px solid #00838f; }
.track-semantics { border-left:3px solid #2e7d32; }
.track-social    { border-left:3px solid #c2185b; }
.track-systems   { border-left:3px solid #512da8; }
.track-security  { border-left:3px solid #c62828; }
.track-mining    { border-left:3px solid #558b2f; }
.track-recsys    { border-left:3px solid #283593; }
.track-industry  { border-left:3px solid #616161; }
.track-history   { border-left:3px solid #8d6e63; }
.track-web4good  { border-left:3px solid #00796b; }
.track-respweb   { border-left:3px solid #e64a19; }
.track-phd       { border-left:3px solid #689f38; }
.track-web4all   { border-left:3px solid #0277bd; }
.track-demos     { border-left:3px solid #3949ab; }
.track-short     { border-left:3px solid #8e24aa; }
.track-best      { border-left:3px solid #f9a825; }
.track-other     { border-left:3px solid #9e9e9e; }
/* Papers toggle button */
.papers-toggle {
  appearance: none; display: inline-flex; align-items: center; justify-content: center; gap: 6px;
  height: 28px; box-sizing: border-box; background: none; border: 1px solid #e0e0e0; border-radius: 5px;
  padding: 0 10px; cursor: pointer; font-size: 0.78rem; color: #555; line-height: 1;
  transition: all .2s; align-self: flex-start; margin: 0;
}
.papers-toggle:hover { background: #f5f0ff; color: #6a00ff; border-color: #6a00ff; }
.papers-toggle[aria-expanded="true"] .toggle-icon { transform: rotate(180deg); }
.papers-count { font-weight: 600; }
.toggle-icon { transition: transform .2s; font-size: 0.7rem; }
/* Papers list */
.papers-container {
  margin-top: 8px; border-top: 1px solid #f0e8ff; padding: 10px 0 0;
  background: linear-gradient(#fcfbff, #fff 55%);
}
.paper-list { list-style: none; margin: 0; padding: 0; }
.paper-item { padding: 7px 0; border-bottom: 1px solid #f5f5f5; display: flex; flex-direction: column; gap: 2px; }
.paper-item:last-child { border-bottom: none; }
.paper-title { font-size: 0.82rem; font-weight: 600; color: #1a1a1a; text-decoration: none; line-height: 1.35; }
.paper-title:hover { color: #6a00ff; text-decoration: underline; }
.acm-link-placeholder { color: #1a1a1a; pointer-events: none; cursor: default; }
.acm-link-placeholder[href]:not([href="#"]) { color: #1565c0; pointer-events: auto; cursor: pointer; }
.paper-authors { font-size: 0.76rem; color: #666; line-height: 1.3; }
/* Timezone note */
.timezone-note {
  background: #f5f5f5; border: 1px solid #e0e0e0; border-radius: 6px;
  padding: 10px 16px; margin-bottom: 24px; font-size: 0.88rem; color: #555;
  display: flex; align-items: center; gap: 8px;
}
.day-heading {
  font-size: 1.3rem; font-weight: 700; color: #1a1a1a;
  padding: 14px 0 10px; border-bottom: 2px solid #6a00ff; margin-bottom: 20px;
}
</style>
'''

COMMON_SCRIPTS = '''
    <script src="../js/header-loader.js"></script>
    <script src="../js/footer-loader.js"></script>
    <script>
    document.addEventListener('DOMContentLoaded', function() {
        document.querySelectorAll('.day-tab-btn').forEach(function(btn) {
            btn.addEventListener('click', function() {
                var target = this.dataset.target;
                document.querySelectorAll('.day-tab-btn').forEach(b => b.classList.remove('active'));
                document.querySelectorAll('.day-panel').forEach(p => p.classList.remove('active'));
                this.classList.add('active');
                document.getElementById(target).classList.add('active');
            });
        });
        var firstBtn = document.querySelector('.day-tab-btn');
        if (firstBtn) firstBtn.click();
    });
    </script>
'''

COLLAPSE_SCRIPT = '''
    <script>
    document.addEventListener('DOMContentLoaded', function() {
        document.querySelectorAll('[data-bs-toggle="collapse"]').forEach(function(btn) {
            btn.addEventListener('click', function() {
                var targetId = this.dataset.bsTarget.replace('#','');
                var panel = document.getElementById(targetId);
                if (!panel) return;
                var expanded = this.getAttribute('aria-expanded') === 'true';
                if (expanded) {
                    panel.classList.remove('show');
                    this.setAttribute('aria-expanded','false');
                } else {
                    panel.classList.add('show');
                    this.setAttribute('aria-expanded','true');
                }
            });
        });
    });
    </script>
'''

COLLAPSE_CSS = '''<style>
.collapse { display: none; }
.collapse.show { display: block; }
</style>
'''

# ─────────────────────────────────────────────
# 5. Build overview.html
# ─────────────────────────────────────────────

def build_overview():
    tabs_html = ''
    panels_html = ''
    for di, day in enumerate(program):
        did = day_id(day)
        tabs_html += f'<button class="day-tab-btn" data-target="{did}">{format_day_tab(day)}</button>\n'
        base_idx = sum(len(program[k]['slots']) for k in range(di))
        slots_html = ''
        for si, slot in enumerate(day['slots']):
            sidx = base_idx + si
            time_html = f'<div class="time-label"><i class="fas fa-clock"></i>{esc(slot["time"])}</div>'
            if slot['special']:
                event = slot['event']
                block = render_keynote_block_overview(event, sidx) if keynote_lookup(event) else f'<div class="slot-special-block {special_class(event)}">{esc(event)}</div>'
                slots_html += f'<div class="time-slot">\n  {time_html}\n  {block}\n</div>\n'
            elif slot['poster']:
                slots_html += f'<div class="time-slot">\n  {time_html}\n  {render_poster_session(slot, False, sidx)}\n</div>\n'
            else:
                cards = ''.join(render_session_card_overview(s) + '\n' for s in (slot['sessions'] or []))
                slots_html += f'<div class="time-slot">\n  {time_html}\n  <div class="session-grid">{cards}</div>\n</div>\n'
        panels_html += f'<div class="day-panel" id="{did}">\n  <h2 class="day-heading">{esc(format_day_heading(day))}</h2>\n  {slots_html}\n</div>\n'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Program Overview - The Web Conference 2026</title>
    {COMMON_HEAD_LINKS}
    {PROGRAM_CSS}
</head>
<body>
    <div id="header-placeholder">Header loading...</div>
    <section class="page-header"><div class="container"><h1>Program Overview</h1></div></section>
    <section class="page-content program-page"><div class="container">
        <div class="timezone-note"><i class="fas fa-info-circle"></i> All times are in <strong>UAE time (GST, UTC+4)</strong>.</div>
        <div class="day-tabs">{tabs_html}</div>
        {panels_html}
    </div></section>
    <div id="footer-placeholder"></div>
    {COMMON_SCRIPTS}
</body>
</html>'''


# ─────────────────────────────────────────────
# 6. Build full-schedule.html
# ─────────────────────────────────────────────

def build_full_schedule():
    tabs_html = ''
    panels_html = ''
    for di, day in enumerate(program):
        did = day_id(day)
        tabs_html += f'<button class="day-tab-btn" data-target="{did}">{format_day_tab(day)}</button>\n'
        base_idx = sum(len(program[k]['slots']) for k in range(di))
        slots_html = ''
        for si, slot in enumerate(day['slots']):
            sidx = base_idx + si
            time_html = f'<div class="time-label"><i class="fas fa-clock"></i>{esc(slot["time"])}</div>'
            if slot['special']:
                event = slot['event']
                block = render_keynote_block_full(event, day, slot, sidx) if keynote_lookup(event) else render_special_block_full(event, day, slot, sidx)
                slots_html += f'<div class="time-slot">\n  {time_html}\n  {block}\n</div>\n'
            elif slot['poster']:
                slots_html += f'<div class="time-slot">\n  {time_html}\n  {render_poster_session(slot, True, sidx, day)}\n</div>\n'
            else:
                cards = ''.join(render_session_card_full(s, day, slot, sidx, si2) + '\n' for si2, s in enumerate(slot['sessions'] or []))
                slots_html += f'<div class="time-slot">\n  {time_html}\n  <div class="session-grid">{cards}</div>\n</div>\n'
        panels_html += f'<div class="day-panel" id="{did}">\n  <h2 class="day-heading">{esc(format_day_heading(day))}</h2>\n  {slots_html}\n</div>\n'

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Full Schedule - The Web Conference 2026</title>
    {COMMON_HEAD_LINKS}
    {PROGRAM_CSS}
    {COLLAPSE_CSS}
</head>
<body>
    <div id="header-placeholder">Header loading...</div>
    <section class="page-header"><div class="container"><h1>Full Schedule</h1></div></section>
    <section class="page-content program-page"><div class="container">
        <div class="timezone-note"><i class="fas fa-info-circle"></i> All times are in <strong>UAE time (GST, UTC+4)</strong>. Click <strong>details</strong> on any session to expand.</div>
        <div class="day-tabs">{tabs_html}</div>
        {panels_html}
    </div></section>
    <div id="footer-placeholder"></div>
    {COMMON_SCRIPTS}
    {COLLAPSE_SCRIPT}
</body>
</html>'''


# ─────────────────────────────────────────────
# 7. Write files
# ─────────────────────────────────────────────

with open('program/overview.html', 'w', encoding='utf-8') as f:
    f.write(build_overview())
print('Written: program/overview.html')

with open('program/full-schedule.html', 'w', encoding='utf-8') as f:
    f.write(build_full_schedule())
print('Written: program/full-schedule.html')

total_papers = sum(
    len(sess.get('papers', []))
    for day in program
    for slot in day['slots']
    for sess in (slot.get('sessions') or [])
)
poster_slots = sum(1 for day in program for slot in day['slots'] if slot.get('poster'))
print(f'\nDays: {len(program)}')
print(f'Poster sessions parsed: {poster_slots}')
print(f'Total papers/items: {total_papers}')
